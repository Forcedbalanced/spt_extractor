"""
SPT Masa PPN Extractor - Full Coverage
Extracts ALL fields from DJP SPT Masa PPN PDF documents.
Usage: python3 ppn_extractor.py <path_to_pdf>
       python3 ppn_extractor.py <folder> [--output results.xlsx]
"""

import re
import sys
import time
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Helpers ────────────────────────────────────────────────────────────────────

def parse_amount(text):
    if not text or str(text).strip() in ('-', '', 'None'):
        return None
    clean = str(text).strip().replace('.', '').replace(',', '.')
    try:
        return float(clean)
    except ValueError:
        return None


def last_number(line):
    """Return the last parseable number on a line, or None."""
    nums = re.findall(r'-?[\d\.]+(?:,\d+)?', line)
    for n in reversed(nums):
        v = parse_amount(n)
        if v is not None:
            return v
    return None


def numbers_on_line(line, min_val=0):
    nums = re.findall(r'-?[\d\.]+(?:,\d+)?', line)
    return [v for n in nums for v in [parse_amount(n)] if v is not None and v >= min_val]


# ── Core extractor ─────────────────────────────────────────────────────────────

def process_spt_ppn(pdf_path):
    pdf_path = Path(pdf_path)
    all_text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                all_text += text + "\n"

    r = {
        "file":                     pdf_path.name,
        "masa_pajak":               None,
        "normal_pembetulan":        None,
        "nama_pkp":                 None,
        "npwp":                     None,
        "tanggal_lapor":            None,

        # ── SECTION I: PENYERAHAN ─────────────────────────────────────────────
        # I.A — terutang PPN
        "ia1_ekspor_dpp":           None,   # I.A.1 Harga Jual/Ekspor
        "ia2_harga_jual":           None,   # I.A.2 Harga Jual
        "ia2_dpp_nilai_lain":       None,   # I.A.2 DPP Nilai Lain (kode 04/05)
        "ia2_ppn":                  None,
        "ia2_ppnbm":                None,
        "ia3_dpp":                  None,   # I.A.3 Turis Pasal 16E (kode 06)
        "ia3_ppn":                  None,
        "ia3_ppnbm":                None,
        "ia4_dpp":                  None,   # I.A.4 Lainnya (kode 01/09/10)
        "ia4_ppn":                  None,
        "ia4_ppnbm":                None,
        "ia5_harga_jual":           None,   # I.A.5 Harga Jual
        "ia5_dpp":                  None,   # I.A.5 Digunggung
        "ia5_ppn":                  None,
        "ia5_ppnbm":                None,
        "ia6_dpp":                  None,   # I.A.6 Pemungut PPN (kode 02/03)
        "ia6_ppn":                  None,
        "ia6_ppnbm":                None,
        "ia7_dpp":                  None,   # I.A.7 Tidak Dipungut (kode 07)
        "ia7_ppn":                  None,
        "ia7_ppnbm":                None,
        "ia8_dpp":                  None,   # I.A.8 Dibebaskan (kode 08)
        "ia8_ppn":                  None,
        "ia8_ppnbm":                None,
        "ia9_dpp":                  None,   # I.A.9 Digunggung fasilitas
        "ia9_ppn":                  None,
        "ia9_ppnbm":                None,
        "ia_jumlah_harga_jual":     None,   # Jumlah I.A (Harga Jual)
        "ia_jumlah_dpp":            None,   # Jumlah I.A DPP
        "ia_jumlah_ppn":            None,   # Jumlah I.A PPN
        "ia_jumlah_ppnbm":          None,   # Jumlah I.A PPnBM
        "ib_tidak_terutang":        None,   # I.B tidak terutang PPN
        "ic_jumlah_penyerahan":     None,   # I.C total

        # ── SECTION II: PEROLEHAN ─────────────────────────────────────────────
        "iia_dpp":                  None,   # II.A Impor/Pemanfaatan luar negeri
        "iia_ppn":                  None,
        "iia_ppnbm":                None,
        "iib_harga_jual":           None,   # II.B DPP Nilai Lain (kode 04/05)
        "iib_dpp_nilai_lain":       None,
        "iib_ppn":                  None,
        "iib_ppnbm":                None,
        "iic_dpp":                  None,   # II.C Selain DPP Nilai Lain (kode 01/09/10)
        "iic_ppn":                  None,
        "iic_ppnbm":                None,
        "iid_dpp":                  None,   # II.D Pemungut PPN (kode 02/03)
        "iid_ppn":                  None,
        "iid_ppnbm":                None,
        "iie_kompensasi":           None,   # II.E Kompensasi kelebihan PM
        "iif_penghitungan_kembali": None,   # II.F Hasil penghitungan kembali
        "iig_jumlah_dpp":           None,   # II.G Jumlah PM diperhitungkan (DPP)
        "iig_jumlah_ppn":           None,   # II.G PPN
        "iih_dpp":                  None,   # II.H tidak dikreditkan/fasilitas
        "iih_ppn":                  None,
        "iih_ppnbm":                None,
        "iii_dpp":                  None,   # II.I digunggung
        "iij_jumlah_dpp":           None,   # II.J total perolehan

        # ── SECTION III: PERHITUNGAN PPN ──────────────────────────────────────
        "iii_a":                    None,   # III.A Pajak Keluaran
        "iii_b":                    None,   # III.B PPN disetor dimuka
        "iii_c":                    None,   # III.C Pajak Masukan (II.G)
        "iii_d":                    None,   # III.D Kelebihan pemungut
        "iii_e":                    None,   # III.E Kurang/(Lebih) bayar
        "iii_f":                    None,   # III.F SPT dibetulkan
        "iii_g":                    None,   # III.G Pembetulan
        "iii_h_status":             None,   # III.H dikompensasi/dikembalikan

        # ── SECTION IV ────────────────────────────────────────────────────────
        "iv_dpp":                   None,   # IV PPN membangun sendiri DPP
        "iv_ppn":                   None,   # IV PPN

        # ── SECTION V ─────────────────────────────────────────────────────────
        "v_ppn_bayar_kembali":      None,   # V Pembayaran kembali PM

        # ── SECTION VI: PPnBM ─────────────────────────────────────────────────
        "via_ppnbm":                None,   # VI.A PPnBM dipungut sendiri
        "vib_ppnbm":                None,   # VI.B Kelebihan pemungutan
        "vic_ppnbm":                None,   # VI.C Kurang/(Lebih) bayar
        "vid_ppnbm":                None,   # VI.D SPT dibetulkan
        "vie_ppnbm":                None,   # VI.E Pembetulan

        # ── SECTION VII: Pemungut PPN ─────────────────────────────────────────
        "viia_dpp":                 None,
        "viia_ppn":                 None,
        "viia_ppnbm":               None,

        # ── SECTION VIII: Pihak Lain ──────────────────────────────────────────
        "viiia_dpp":                None,
        "viiia_ppn":                None,
        "viiia_ppnbm":              None,
    }

    lines = all_text.split('\n')

    # ── HEADER ────────────────────────────────────────────────────────────────
    for line in lines:
        m = re.search(
            r'(Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|September'
            r'|Oktober|November|Desember)\s+(\d{4})\s+\d+\s+s\.d\s+\d+\s+'
            r'(NORMAL|PEMBETULAN(?:\s+KE-\d+)?)',
            line, re.IGNORECASE)
        if m:
            r["masa_pajak"] = f"{m.group(1)} {m.group(2)}"
            r["normal_pembetulan"] = m.group(3).strip()

        m = re.search(r'NAMA PKP\s*:\s*(.+?)\s+NPWP\s*:', line)
        if m:
            r["nama_pkp"] = m.group(1).strip()

        m = re.search(r'NPWP\s*:\s*(\d{15,16})', line)
        if m:
            r["npwp"] = m.group(1).strip()

        m = re.search(
            r',\s*(\d{1,2})\s+(Januari|Februari|Maret|April|Mei|Juni|Juli'
            r'|Agustus|September|Oktober|November|Desember)\s+(\d{4})\s*$',
            line, re.IGNORECASE)
        if m and r["tanggal_lapor"] is None:
            r["tanggal_lapor"] = f"{m.group(1)} {m.group(2)} {m.group(3)}"

    # ── SECTION I: PENYERAHAN ─────────────────────────────────────────────────
    # Items 1–9 are numbered lines; grab numbers by position.
    # Items 7-9 start with "Penyerahan yang mendapat fasilitas", so match
    # any numbered line starting with "Ekspor" or "Penyerahan".
    item_lines = {}
    for line in lines:
        m = re.match(r'^\s*([1-9])\.\s+(Ekspor|Penyerahan)', line, re.IGNORECASE)
        if m:
            item_lines[int(m.group(1))] = line

    def pick(nums, idx, default=None):
        return nums[idx] if idx < len(nums) else default

    def strip_item_num(line):
        """Remove leading item number like '2. ' before parsing numbers."""
        return re.sub(r'^\s*\d+\.\s+', '', line)

    # I.A.1 — only Harga Jual column
    if 1 in item_lines:
        ns = numbers_on_line(strip_item_num(item_lines[1]))
        r["ia1_ekspor_dpp"] = pick(ns, 0, 0)

    # I.A.2 — Harga Jual | DPP Nilai Lain | PPN | PPnBM
    if 2 in item_lines:
        ns = numbers_on_line(strip_item_num(item_lines[2]))
        r["ia2_harga_jual"] = pick(ns, 0, 0)
        r["ia2_dpp_nilai_lain"] = pick(ns, 1, pick(ns, 0, 0))
        r["ia2_ppn"] = pick(ns, 2, pick(ns, 1, 0))
        r["ia2_ppnbm"] = pick(ns, 3, 0)

    # I.A.3 — Harga Jual | DPP | PPN | PPnBM
    if 3 in item_lines:
        ns = numbers_on_line(strip_item_num(item_lines[3]))
        r["ia3_dpp"] = pick(ns, 0, 0)
        r["ia3_ppn"] = pick(ns, 1, 0)   # actually index 2 but 0-cols collapse when 0
        r["ia3_ppnbm"] = pick(ns, 3, 0)

    # I.A.4
    if 4 in item_lines:
        ns = numbers_on_line(strip_item_num(item_lines[4]))
        r["ia4_dpp"] = pick(ns, 0, 0)
        r["ia4_ppn"] = pick(ns, 1, 0)
        r["ia4_ppnbm"] = pick(ns, 2, 0)

    # I.A.5
    if 5 in item_lines:
        ns = numbers_on_line(strip_item_num(item_lines[5]))
        r["ia5_harga_jual"] = pick(ns, 0, 0)
        r["ia5_dpp"] = pick(ns, 1, 0)
        r["ia5_ppn"] = pick(ns, 2, 0)
        r["ia5_ppnbm"] = pick(ns, 3, 0)

    # I.A.6
    if 6 in item_lines:
        ns = numbers_on_line(strip_item_num(item_lines[6]))
        r["ia6_dpp"] = pick(ns, 0, 0)
        r["ia6_ppn"] = pick(ns, 1, 0)
        r["ia6_ppnbm"] = pick(ns, 2, 0)

    # I.A.7
    if 7 in item_lines:
        ns = numbers_on_line(strip_item_num(item_lines[7]))
        r["ia7_dpp"] = pick(ns, 0, 0)
        r["ia7_ppn"] = pick(ns, 1, 0)
        r["ia7_ppnbm"] = pick(ns, 2, 0)

    # I.A.8
    if 8 in item_lines:
        ns = numbers_on_line(strip_item_num(item_lines[8]))
        r["ia8_dpp"] = pick(ns, 0, 0)
        r["ia8_ppn"] = pick(ns, 1, 0)
        r["ia8_ppnbm"] = pick(ns, 2, 0)

    # I.A.9
    if 9 in item_lines:
        ns = numbers_on_line(strip_item_num(item_lines[9]))
        r["ia9_dpp"] = pick(ns, 0, 0)
        r["ia9_ppn"] = pick(ns, 1, 0)
        r["ia9_ppnbm"] = pick(ns, 2, 0)

    for line in lines:
        if re.search(r'Jumlah \(I\.A\.1 \+ I\.A\.2', line, re.IGNORECASE):
            after = re.sub(r'^.*?\)', '', line)
            ns = numbers_on_line(after)
            r["ia_jumlah_harga_jual"] = pick(ns, 0, 0)
            if len(ns) >= 4:
                # All 4 columns present: harga_jual, dpp, ppn, ppnbm
                r["ia_jumlah_dpp"] = pick(ns, 1, 0)
                r["ia_jumlah_ppn"] = pick(ns, 2, 0)
                r["ia_jumlah_ppnbm"] = pick(ns, 3, 0)
            else:
                # pdfplumber drops the DPP column on this row; 3 values are
                # harga_jual, ppn, ppnbm — compute dpp from individual items below
                r["ia_jumlah_dpp"] = None
                r["ia_jumlah_ppn"] = pick(ns, 1, 0)
                r["ia_jumlah_ppnbm"] = pick(ns, 2, 0)
            break

    # If DPP Jumlah I.A was not extractable from the row, derive it from items.
    # Row 1 (Ekspor) has no DPP column, so excluded from this sum.
    if r["ia_jumlah_dpp"] is None:
        dpp_items = [
            r["ia2_harga_jual"], r["ia3_dpp"], r["ia4_dpp"],
            r["ia5_dpp"], r["ia6_dpp"], r["ia7_dpp"], r["ia8_dpp"], r["ia9_dpp"],
        ]
        r["ia_jumlah_dpp"] = sum(v for v in dpp_items if v is not None)

    for line in lines:
        if re.search(r'Penyerahan barang.jasa yang tidak terutang PPN', line, re.IGNORECASE):
            r["ib_tidak_terutang"] = last_number(line) or 0
            break

    for line in lines:
        if re.search(r'Jumlah seluruh penyerahan', line, re.IGNORECASE):
            r["ic_jumlah_penyerahan"] = last_number(line) or 0
            break

    # ── SECTION II: PEROLEHAN ─────────────────────────────────────────────────
    for i, line in enumerate(lines):
        combined = line + (" " + lines[i+1] if i+1 < len(lines) else "")

        if re.search(r'II\.\s*A\.|Impor BKP.*Pemanfaatan BKP Tidak Berwujud', line, re.IGNORECASE):
            ns = numbers_on_line(combined)
            r["iia_dpp"] = pick(ns, 0, 0)
            r["iia_ppn"] = pick(ns, 1, 0)
            r["iia_ppnbm"] = pick(ns, 2, 0)

        elif re.search(
                r'Perolehan BKP.JKP dari dalam negeri dengan DPP Nilai Lain',
                line, re.IGNORECASE):
            ns = numbers_on_line(combined)
            r["iib_harga_jual"] = pick(ns, 0, 0)
            r["iib_dpp_nilai_lain"] = pick(ns, 1, 0)
            r["iib_ppn"] = pick(ns, 2, 0)
            r["iib_ppnbm"] = pick(ns, 3, 0)

        elif re.search(
                r'Perolehan BKP.JKP dari dalam negeri selain dengan DPP Nilai Lain',
                line, re.IGNORECASE):
            ns = numbers_on_line(combined)
            r["iic_dpp"] = pick(ns, 0, 0)
            r["iic_ppn"] = pick(ns, 1, 0)
            r["iic_ppnbm"] = pick(ns, 2, 0)

        elif re.search(
                r'Perolehan BKP.JKP dari dalam negeri sebagai Pemungut PPN',
                line, re.IGNORECASE):
            ns = numbers_on_line(combined)
            r["iid_dpp"] = pick(ns, 0, 0)
            r["iid_ppn"] = pick(ns, 1, 0)
            r["iid_ppnbm"] = pick(ns, 2, 0)

        elif re.search(r'Kompensasi kelebihan Pajak Masukan', line, re.IGNORECASE):
            r["iie_kompensasi"] = last_number(line) or 0

        elif (re.search(r'Hasil penghitungan kembali Pajak Masukan', line, re.IGNORECASE)
              and not re.match(r'^\s*\d+\.', line)):
            # Exclude Section IX numbered list items (e.g. "2. Hasil Penghitungan...")
            # which match the same words but hold only the item number as a value.
            v = last_number(line)
            r["iif_penghitungan_kembali"] = v if v is not None else 0

        elif re.search(
                r'Jumlah Pajak Masukan yang dapat diperhitungkan',
                line, re.IGNORECASE):
            after = re.sub(r'^.*?\)', '', line)
            ns = numbers_on_line(after)
            r["iig_jumlah_dpp"] = pick(ns, 0, 0)
            r["iig_jumlah_ppn"] = pick(ns, 1, pick(ns, 0, 0))

        elif re.search(
                r'Impor atau perolehan BKP.JKP yang Pajak Masukannya tidak dikreditkan',
                line, re.IGNORECASE):
            ns = numbers_on_line(combined)
            r["iih_dpp"] = pick(ns, 0, 0)
            r["iih_ppn"] = pick(ns, 1, 0)
            r["iih_ppnbm"] = pick(ns, 2, 0)

        elif re.search(
                r'Impor atau perolehan BKP.JKP dengan Faktur Pajak'
                r' yang dilaporkan secara digunggung',
                line, re.IGNORECASE):
            r["iii_dpp"] = last_number(line) or 0

        elif re.search(r'Jumlah perolehan \(II\.A \+ II\.B', line, re.IGNORECASE):
            after = re.sub(r'^.*?\)', '', line)
            r["iij_jumlah_dpp"] = last_number(after) or 0

    # ── SECTION III: PERHITUNGAN PPN ──────────────────────────────────────────
    for line in lines:
        if re.search(r'Pajak Keluaran yang harus dipungut sendiri \(I\.A\.2', line, re.IGNORECASE):
            r["iii_a"] = last_number(line) or 0

        elif re.search(r'PPN disetor dimuka dalam Masa Pajak yang sama', line, re.IGNORECASE):
            r["iii_b"] = last_number(line) or 0

        elif re.search(r'Pajak Masukan yang dapat diperhitungkan \(II\.G\)', line, re.IGNORECASE):
            r["iii_c"] = last_number(line) or 0

        elif (re.search(r'Kelebihan pemungutan PPN oleh Pemungut PPN', line, re.IGNORECASE)
              and 'III' not in line[:10]):
            r["iii_d"] = last_number(line) or 0

        elif re.search(r'PPN kurang atau \(lebih\) bayar \(III\.A', line, re.IGNORECASE):
            r["iii_e"] = last_number(line)

        elif re.search(
                r'PPN kurang atau \(lebih\) bayar pada SPT yang dibetulkan sebelumnya',
                line, re.IGNORECASE):
            r["iii_f"] = last_number(line)

        elif re.search(
                r'PPN kurang atau \(lebih\) bayar karena pembetulan SPT \(III\.E',
                line, re.IGNORECASE):
            r["iii_g"] = last_number(line)

        elif re.search(r'diminta untuk.*dikompensasikan', line, re.IGNORECASE):
            if re.search(r'x\s*1\.', line):
                r["iii_h_status"] = "dikompensasikan"
            elif re.search(r'x\s*2\.', line):
                r["iii_h_status"] = "dikembalikan pendahuluan"
            elif re.search(r'x\s*3\.', line):
                r["iii_h_status"] = "dikembalikan pemeriksaan"

    # ── SECTION IV ────────────────────────────────────────────────────────────
    for line in lines:
        if re.search(r'PPN Terutang\s+\d', line, re.IGNORECASE):
            ns = numbers_on_line(line)
            r["iv_dpp"] = pick(ns, 0, 0)
            r["iv_ppn"] = pick(ns, 1, 0)
            break

    # ── SECTION V ─────────────────────────────────────────────────────────────
    for line in lines:
        if re.search(r'PPN Yang Wajib Dibayar Kembali', line, re.IGNORECASE):
            r["v_ppn_bayar_kembali"] = last_number(line) or 0
            break

    # ── SECTION VI: PPnBM ─────────────────────────────────────────────────────
    for line in lines:
        if re.search(r'PPnBM yang harus dipungut sendiri \(I\.A\.2', line, re.IGNORECASE):
            r["via_ppnbm"] = last_number(line) or 0
        elif re.search(r'Kelebihan pemungutan PPnBM oleh Pemungut PPN', line, re.IGNORECASE):
            r["vib_ppnbm"] = last_number(line) or 0
        elif re.search(r'PPnBM kurang atau \(lebih\) bayar \(VI\.A', line, re.IGNORECASE):
            r["vic_ppnbm"] = last_number(line)
        elif re.search(
                r'PPnBM kurang atau \(lebih\) bayar pada SPT yang dibetulkan',
                line, re.IGNORECASE):
            r["vid_ppnbm"] = last_number(line)
        elif re.search(
                r'PPnBM kurang atau \(lebih\) bayar karena pembetulan SPT \(VI\.C',
                line, re.IGNORECASE):
            r["vie_ppnbm"] = last_number(line)

    # ── SECTION VII ───────────────────────────────────────────────────────────
    for line in lines:
        if re.search(r'VII.*Jumlah PPN dan PPnBM yang dipungut', line, re.IGNORECASE):
            ns = numbers_on_line(line)
            r["viia_dpp"] = pick(ns, 0, 0)
            r["viia_ppn"] = pick(ns, 1, 0)
            r["viia_ppnbm"] = pick(ns, 2, 0)
            break

    # ── SECTION VIII ──────────────────────────────────────────────────────────
    for line in lines:
        if re.search(r'VIII.*Jumlah PPN dan PPnBM yang dipungut', line, re.IGNORECASE):
            ns = numbers_on_line(line)
            r["viiia_dpp"] = pick(ns, 0, 0)
            r["viiia_ppn"] = pick(ns, 1, 0)
            r["viiia_ppnbm"] = pick(ns, 2, 0)
            break

    return r


# ── Bulk processor ─────────────────────────────────────────────────────────────

def process_folder(folder_path):
    folder = Path(folder_path)
    pdf_files = list(folder.glob("*.pdf")) + list(folder.glob("*.PDF"))
    pdf_files = list(dict.fromkeys(pdf_files))

    def sort_key(p):
        m = re.match(r'^(\d+)', p.stem.strip())
        return int(m.group(1)) if m else 999

    pdf_files = sorted(pdf_files, key=sort_key)

    if not pdf_files:
        print(f"No PDF files found in: {folder}")
        return []

    t_start = time.time()
    results = []
    for i, pdf_file in enumerate(pdf_files, 1):
        print(f"[{i}/{len(pdf_files)}] Processing: {pdf_file.name}", end=" ... ")
        if not pdf_file.is_file():
            print("SKIP (not a file)")
            continue
        try:
            data = process_spt_ppn(pdf_file)
            results.append(data)
            print("OK")
        except Exception as e:
            print(f"ERROR: {e}")
            results.append({"file": pdf_file.name, "error": str(e)})

    elapsed = time.time() - t_start
    avg = elapsed / len(results)
    print(f"\nSelesai: {len(results)} file diproses dalam {elapsed:.2f} detik "
          f"({avg:.2f} detik/file rata-rata)")
    return results


# ── Excel column definitions ───────────────────────────────────────────────────
# (label, key, is_number)

# Exactly 28 columns matching the reference Excel format
COLUMNS = [
    # INFO (cols 1-5)
    ("File",                        "file",                     False),
    ("Masa Pajak",                  "masa_pajak",               False),
    ("Status SPT",                  "normal_pembetulan",        False),
    ("Nama PKP",                    "nama_pkp",                 False),
    ("NPWP",                        "npwp",                     False),
    # Section I (cols 6-16)
    ("I.A.1 Ekspor BKP/JKP",        "ia1_ekspor_dpp",           True),
    ("I.A.2 DPP",                   "ia2_harga_jual",           True),
    ("I.A.2 PPN",                   "ia2_ppn",                  True),
    ("I.A.3 DPP",                   "ia3_dpp",                  True),
    ("I.A.3 PPN",                   "ia3_ppn",                  True),
    ("I.A.4 DPP",                   "ia4_dpp",                  True),
    ("I.A.4 PPN",                   "ia4_ppn",                  True),
    ("I.A.5 Harga Jual",            "ia5_harga_jual",           True),
    ("Jumlah I.A DPP",              "ia_jumlah_dpp",            True),
    ("Jumlah I.A PPN",              "ia_jumlah_ppn",            True),
    ("Total Penyerahan (I.C)",      "ic_jumlah_penyerahan",     True),
    # Section II (cols 15-19)
    ("II.B DPP",                    "iib_harga_jual",           True),
    ("II.B PPN",                    "iib_ppn",                  True),
    ("II.G DPP",                    "iig_jumlah_dpp",           True),
    ("II.G PPN (Pajak Masukan)",    "iig_jumlah_ppn",           True),
    ("Total Perolehan (II.J)",      "iij_jumlah_dpp",           True),
    # Section III (cols 20-28)
    ("III.A Pajak Keluaran",        "iii_a",                    True),
    ("III.B PPN Disetor Dimuka",    "iii_b",                    True),
    ("III.C Pajak Masukan",         "iii_c",                    True),
    ("III.D Kelebihan Pemungut",    "iii_d",                    True),
    ("III.E Kurang/(Lebih) Bayar",  "iii_e",                    True),
    ("III.F SPT Dibetulkan",        "iii_f",                    True),
    ("III.G Pembetulan",            "iii_g",                    True),
    ("III.H Status",                "iii_h_status",             False),
    ("Tanggal Lapor",               "tanggal_lapor",            False),
]

SECTION_LABELS = [
    ("INFO WAJIB PAJAK",                            1,   5),
    ("I. PENYERAHAN BARANG DAN JASA",               6,  16),
    ("II. PEROLEHAN BARANG DAN JASA",              17,  21),
    ("III. PERHITUNGAN PPN KURANG/LEBIH BAYAR",    22,  30),
]

SECTION_COLORS = ["1F4E79", "156082", "0E7C7B", "375623"]

NUMBER_FMT = '#,##0;(#,##0);"-"'


def export_to_excel(results, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "SPT Masa PPN"

    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_color = {}
    for (label, sc, ec), color in zip(SECTION_LABELS, SECTION_COLORS):
        sl = get_column_letter(sc)
        el = get_column_letter(ec)
        ws.merge_cells(f"{sl}1:{el}1")
        cell = ws[f"{sl}1"]
        cell.value = label
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", start_color=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        for c in range(sc, ec + 1):
            col_color[c] = color

    ws.row_dimensions[1].height = 24

    for col_idx, (header, key, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", start_color=col_color.get(col_idx, "1F4E79"))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[2].height = 36

    for row_idx, record in enumerate(results, 3):
        is_error = "error" in record
        alt = (row_idx % 2 == 0)

        for col_idx, (header, key, is_num) in enumerate(COLUMNS, 1):
            value = record.get(key)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = Font(name="Arial", size=9)

            if is_num and value is not None and not is_error:
                cell.number_format = NUMBER_FMT
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

            if is_error:
                cell.fill = PatternFill("solid", start_color="FCE4D6")
            elif alt:
                cell.fill = PatternFill("solid", start_color="EBF3FB")

            cell.border = border

        ws.row_dimensions[row_idx].height = 18

    # Column widths matching reference exactly
    wide = {1: 30, 2: 16, 3: 12, 4: 35, 5: 20, 6: 18, 7: 18, 8: 18}
    for col_idx in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = wide.get(col_idx, 18)

    ws.freeze_panes = "F3"

    # Summary sheet
    ws2 = wb.create_sheet("Ringkasan")
    thin2 = Side(style='thin', color='BFBFBF')
    b2 = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)
    for c, h in enumerate(["Keterangan", "Jumlah"], 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center")
        cell.border = b2
    for row_idx, (label, val) in enumerate([
        ("Total file diproses", len(results)),
        ("Berhasil diekstrak", sum(1 for r in results if "error" not in r)),
        ("Gagal diproses",     sum(1 for r in results if "error" in r)),
    ], 2):
        ws2.cell(row=row_idx, column=1, value=label).font = Font(name="Arial")
        ws2.cell(row=row_idx, column=2, value=val).font = Font(name="Arial")
        for c in range(1, 3):
            ws2.cell(row=row_idx, column=c).border = b2
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14

    wb.save(output_path)
    print(f"Exported: {output_path}  ({len(results)} rows, {len(COLUMNS)} kolom)")


# ── CLI ────────────────────────────────────────────────────────────────────────

_DEFAULT_FOLDER = "PPN"

if __name__ == "__main__":
    args = sys.argv[1:]

    output_xlsx = None
    if "--output" in args:
        idx = args.index("--output")
        if idx + 1 < len(args):
            output_xlsx = Path(args[idx + 1])
        args = [a for i, a in enumerate(args) if i != idx and i != idx + 1]

    # Auto-detect PDF folder when no positional argument is given
    if not args:
        target = Path(_DEFAULT_FOLDER)
        if not target.is_dir():
            print(
                f"Error: default folder '{_DEFAULT_FOLDER}' not found. "
                "Pass a folder or file path.")
            sys.exit(1)
    else:
        target = Path(args[0])

    if target.is_dir():
        if output_xlsx is None:
            output_xlsx = Path("hasil_ekstraksi_ppn.xlsx")  # project root
        results = process_folder(target)
        if results:
            export_to_excel(results, output_xlsx)
    elif target.is_file():
        data = process_spt_ppn(target)
        if output_xlsx:
            export_to_excel([data], output_xlsx)
        else:
            for k, v in data.items():
                print(f"{k}: {v}")
    else:
        print(f"Error: '{target}' not found.")
        print("Usage:")
        print(f"  No args : python ppn_extractor.py              (uses ./{_DEFAULT_FOLDER}/)")
        print("  Folder  : python ppn_extractor.py <folder>   [--output out.xlsx]")
        print("  Single  : python ppn_extractor.py <file.pdf> [--output out.xlsx]")
        sys.exit(1)
