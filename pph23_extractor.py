"""
SPT Masa PPh Unifikasi (PPh 23) Extractor
Extracts fields from DJP SPT Masa PPh Unifikasi PDF documents.
Usage: python3 pph23_extractor.py <path_to_pdf>
       python3 pph23_extractor.py <folder> [--output results.xlsx]
"""

import re
import sys
import time
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Helpers ────────────────────────────────────────────────────────────────────

def parse_amount(text):
    if not text or str(text).strip() in ('-', '', 'None'):
        return None
    clean = str(text).strip().replace('.', '').replace(',', '.')
    try:
        return float(clean)
    except ValueError:
        return None

def numbers_on_line(line):
    nums = re.findall(r'-?[\d\.]+(?:,\d+)?', line)
    return [v for n in nums for v in [parse_amount(n)] if v is not None]


# ── Core extractor ─────────────────────────────────────────────────────────────

def process_spt_pph23(pdf_path):
    pdf_path = Path(pdf_path)
    all_text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                all_text += text + "\n"

    r = {
        "file":                         pdf_path.name,

        # ── HEADER ────────────────────────────────────────────────────────────
        "masa_pajak":                   None,   # e.g. "Januari 2025"
        "status_spt":                   None,   # NORMAL / PEMBETULAN
        "npwp":                         None,
        "nama":                         None,
        "alamat":                       None,
        "no_telepon":                   None,

        # ── B. PAJAK PENGHASILAN ───────────────────────────────────────────────
        # PPh Pasal 4 ayat 2
        "pph4a2_setor_sendiri":         None,
        "pph4a2_pemotongan":            None,
        "pph4a2_ditanggung_pemerintah": None,
        "pph4a2_harus_disetor":         None,
        "pph4a2_telah_disetor":         None,
        "pph4a2_kurang_lebih_setor":    None,

        # PPh Pasal 4 ayat 2 – detail KJS
        "kjs_411128_100_pemotongan":    None,
        "kjs_411128_100_harus_disetor": None,
        "kjs_411128_100_telah_disetor": None,
        "kjs_411128_100_selisih":       None,

        "kjs_411128_402_pemotongan":    None,
        "kjs_411128_402_harus_disetor": None,
        "kjs_411128_402_telah_disetor": None,
        "kjs_411128_402_selisih":       None,

        "kjs_411128_403_pemotongan":    None,
        "kjs_411128_403_harus_disetor": None,
        "kjs_411128_403_telah_disetor": None,
        "kjs_411128_403_selisih":       None,

        # PPh Pasal 15
        "pph15_setor_sendiri":          None,
        "pph15_pemotongan":             None,
        "pph15_ditanggung_pemerintah":  None,
        "pph15_harus_disetor":          None,
        "pph15_telah_disetor":          None,
        "pph15_kurang_lebih_setor":     None,

        "kjs_411128_600_pemotongan":    None,
        "kjs_411128_600_harus_disetor": None,
        "kjs_411128_600_telah_disetor": None,
        "kjs_411128_600_selisih":       None,

        "kjs_411129_600_pemotongan":    None,
        "kjs_411129_600_harus_disetor": None,
        "kjs_411129_600_telah_disetor": None,
        "kjs_411129_600_selisih":       None,

        # PPh Pasal 22
        "pph22_setor_sendiri":          None,
        "pph22_pemotongan":             None,
        "pph22_ditanggung_pemerintah":  None,
        "pph22_harus_disetor":          None,
        "pph22_telah_disetor":          None,
        "pph22_kurang_lebih_setor":     None,

        "kjs_411122_100_pemotongan":    None,
        "kjs_411122_100_harus_disetor": None,
        "kjs_411122_100_telah_disetor": None,
        "kjs_411122_100_selisih":       None,

        "kjs_411122_900_pemotongan":    None,
        "kjs_411122_900_harus_disetor": None,
        "kjs_411122_900_telah_disetor": None,
        "kjs_411122_900_selisih":       None,

        "kjs_411122_910_pemotongan":    None,
        "kjs_411122_910_harus_disetor": None,
        "kjs_411122_910_telah_disetor": None,
        "kjs_411122_910_selisih":       None,

        # PPh Pasal 23  ← the main section of interest
        "pph23_setor_sendiri":          None,
        "pph23_pemotongan":             None,
        "pph23_ditanggung_pemerintah":  None,
        "pph23_harus_disetor":          None,
        "pph23_telah_disetor":          None,
        "pph23_kurang_lebih_setor":     None,

        "kjs_411124_100_pemotongan":    None,
        "kjs_411124_100_harus_disetor": None,
        "kjs_411124_100_telah_disetor": None,
        "kjs_411124_100_selisih":       None,

        # PPh Pasal 26
        "pph26_setor_sendiri":          None,
        "pph26_pemotongan":             None,
        "pph26_ditanggung_pemerintah":  None,
        "pph26_harus_disetor":          None,
        "pph26_telah_disetor":          None,
        "pph26_kurang_lebih_setor":     None,

        "kjs_411127_110_pemotongan":    None,
        "kjs_411127_110_harus_disetor": None,
        "kjs_411127_110_telah_disetor": None,
        "kjs_411127_110_selisih":       None,

        # Total
        "total_setor_sendiri":          None,
        "total_pemotongan":             None,
        "total_ditanggung_pemerintah":  None,
        "total_harus_disetor":          None,
        "total_telah_disetor":          None,

        # ── C. TANDA TANGAN ───────────────────────────────────────────────────
        "penandatangan_nama":           None,
        "tanggal_tanda_tangan":         None,
    }

    lines = all_text.split('\n')

    # ── HEADER ────────────────────────────────────────────────────────────────
    for i, line in enumerate(lines):
        # Masa Pajak  (e.g. "Januari 2025")
        m = re.search(r'(Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|September|Oktober|November|Desember)\s+(\d{4})', line, re.IGNORECASE)
        if m and r["masa_pajak"] is None:
            r["masa_pajak"] = f"{m.group(1)} {m.group(2)}"

        # Status SPT
        m = re.search(r'\b(NORMAL|PEMBETULAN(?:\s+KE-\d+)?)\b', line, re.IGNORECASE)
        if m and r["status_spt"] is None:
            r["status_spt"] = m.group(1).strip().upper()

        # NPWP
        m = re.search(r'NPWP[/\w]*\s*[:/]\s*(\d{15,16})', line, re.IGNORECASE)
        if m:
            r["npwp"] = m.group(1)

        # NAMA (A.2)
        m = re.search(r'A\.2\s+NAMA\s*:\s*(.+)', line, re.IGNORECASE)
        if m:
            r["nama"] = m.group(1).strip()

        # ALAMAT (A.3)
        m = re.search(r'A\.3\s+ALAMAT\s*:\s*(.+)', line, re.IGNORECASE)
        if m:
            r["alamat"] = m.group(1).strip()

        # No. Telepon (A.4)
        m = re.search(r'A\.4\s+NO\.\s*TELEPON\s*:\s*(\S+)', line, re.IGNORECASE)
        if m:
            r["no_telepon"] = m.group(1).strip()

        # Penandatangan nama (C.2)
        m = re.search(r'C\.2\s+NAMA\s*:\s*(.+)', line, re.IGNORECASE)
        if m:
            r["penandatangan_nama"] = m.group(1).strip()

        # Tanggal tanda tangan (C.3)
        m = re.search(r'C\.3\s+TANGGAL\s*:\s*(.+)', line, re.IGNORECASE)
        if m:
            r["tanggal_tanda_tangan"] = m.group(1).strip()

    # ── SECTION B: PPh rows ───────────────────────────────────────────────────
    # Each "parent" PPh row (e.g. "PPh Pasal 23") has columns:
    #   B.3 Setor Sendiri | B.4 Pemotongan | B.5 Ditanggung Pemerintah |
    #   B.6 Harus Disetor | B.7 Telah Disetor | B.8 Kurang/Lebih Setor
    #
    # Each KJS sub-row has only:
    #   Pemotongan | Harus Disetor | Telah Disetor | Selisih  (4 values)
    #
    # Strategy: scan for keyword in line, then collect numbers.

    def extract_row_6(line):
        """Return up to 6 numeric values from a parent PPh row.
        Strip a leading single-digit row number (1-6) so it isn't treated as data."""
        clean = re.sub(r'^\s*\d\s+', '', line)
        # Also strip the Pasal number (e.g. '4', '15', '22', '23', '26') that
        # may appear right after the label and before the actual values.
        clean = re.sub(r'\bPasal\s+\d+(?:\s+ayat\s+\d+)?\b', '', clean, flags=re.IGNORECASE)
        nums = numbers_on_line(clean)
        return nums[:6] if len(nums) >= 6 else nums

    def extract_kjs_4(line):
        """Return [pemotongan, harus_disetor, telah_disetor, selisih] from a KJS line.
        KJS lines have 5 values: Setor Sendiri | Pemotongan | Harus Disetor | Telah Disetor | Selisih
        We skip index 0 (Setor Sendiri) and return indices 1-4."""
        clean = re.sub(r'\bKJS\b[^\d]*\d{6}[-\s]\d{3}\b', '', line, flags=re.IGNORECASE)
        nums = numbers_on_line(clean)
        # Skip Setor Sendiri (index 0), return the next 4
        if len(nums) >= 5:
            return nums[1:5]
        elif len(nums) == 4:
            return nums[:4]
        else:
            return nums

    def fill_6(nums, keys):
        for i, k in enumerate(keys):
            if i < len(nums):
                r[k] = nums[i]

    def fill_4(nums, keys):
        for i, k in enumerate(keys):
            if i < len(nums):
                r[k] = nums[i]

    for line in lines:
        # ---- PPh Pasal 4 ayat 2 ----
        if re.search(r'PPh\s+Pasal\s+4\s+ayat\s+2', line, re.IGNORECASE):
            fill_6(extract_row_6(line), [
                "pph4a2_setor_sendiri", "pph4a2_pemotongan",
                "pph4a2_ditanggung_pemerintah", "pph4a2_harus_disetor",
                "pph4a2_telah_disetor", "pph4a2_kurang_lebih_setor",
            ])

        elif re.search(r'KJS\s*[:\s]*411128[-\s]100', line, re.IGNORECASE):
            fill_4(extract_kjs_4(line), [
                "kjs_411128_100_pemotongan", "kjs_411128_100_harus_disetor",
                "kjs_411128_100_telah_disetor", "kjs_411128_100_selisih",
            ])

        elif re.search(r'KJS\s*[:\s]*411128[-\s]402', line, re.IGNORECASE):
            fill_4(extract_kjs_4(line), [
                "kjs_411128_402_pemotongan", "kjs_411128_402_harus_disetor",
                "kjs_411128_402_telah_disetor", "kjs_411128_402_selisih",
            ])

        elif re.search(r'KJS\s*[:\s]*411128[-\s]403', line, re.IGNORECASE):
            fill_4(extract_kjs_4(line), [
                "kjs_411128_403_pemotongan", "kjs_411128_403_harus_disetor",
                "kjs_411128_403_telah_disetor", "kjs_411128_403_selisih",
            ])

        # ---- PPh Pasal 15 ----
        elif re.search(r'PPh\s+Pasal\s+15\b', line, re.IGNORECASE):
            fill_6(extract_row_6(line), [
                "pph15_setor_sendiri", "pph15_pemotongan",
                "pph15_ditanggung_pemerintah", "pph15_harus_disetor",
                "pph15_telah_disetor", "pph15_kurang_lebih_setor",
            ])

        elif re.search(r'KJS\s*[:\s]*411128[-\s]600', line, re.IGNORECASE):
            fill_4(extract_kjs_4(line), [
                "kjs_411128_600_pemotongan", "kjs_411128_600_harus_disetor",
                "kjs_411128_600_telah_disetor", "kjs_411128_600_selisih",
            ])

        elif re.search(r'KJS\s*[:\s]*411129[-\s]600', line, re.IGNORECASE):
            fill_4(extract_kjs_4(line), [
                "kjs_411129_600_pemotongan", "kjs_411129_600_harus_disetor",
                "kjs_411129_600_telah_disetor", "kjs_411129_600_selisih",
            ])

        # ---- PPh Pasal 22 ----
        elif re.search(r'PPh\s+Pasal\s+22\b', line, re.IGNORECASE):
            fill_6(extract_row_6(line), [
                "pph22_setor_sendiri", "pph22_pemotongan",
                "pph22_ditanggung_pemerintah", "pph22_harus_disetor",
                "pph22_telah_disetor", "pph22_kurang_lebih_setor",
            ])

        elif re.search(r'KJS\s*[:\s]*411122[-\s]100', line, re.IGNORECASE):
            fill_4(extract_kjs_4(line), [
                "kjs_411122_100_pemotongan", "kjs_411122_100_harus_disetor",
                "kjs_411122_100_telah_disetor", "kjs_411122_100_selisih",
            ])

        elif re.search(r'KJS\s*[:\s]*411122[-\s]900', line, re.IGNORECASE):
            fill_4(extract_kjs_4(line), [
                "kjs_411122_900_pemotongan", "kjs_411122_900_harus_disetor",
                "kjs_411122_900_telah_disetor", "kjs_411122_900_selisih",
            ])

        elif re.search(r'KJS\s*[:\s]*411122[-\s]910', line, re.IGNORECASE):
            fill_4(extract_kjs_4(line), [
                "kjs_411122_910_pemotongan", "kjs_411122_910_harus_disetor",
                "kjs_411122_910_telah_disetor", "kjs_411122_910_selisih",
            ])

        # ---- PPh Pasal 23 ----
        elif re.search(r'PPh\s+Pasal\s+23\b', line, re.IGNORECASE):
            fill_6(extract_row_6(line), [
                "pph23_setor_sendiri", "pph23_pemotongan",
                "pph23_ditanggung_pemerintah", "pph23_harus_disetor",
                "pph23_telah_disetor", "pph23_kurang_lebih_setor",
            ])

        elif re.search(r'KJS\s*[:\s]*411124[-\s]100', line, re.IGNORECASE):
            fill_4(extract_kjs_4(line), [
                "kjs_411124_100_pemotongan", "kjs_411124_100_harus_disetor",
                "kjs_411124_100_telah_disetor", "kjs_411124_100_selisih",
            ])

        # ---- PPh Pasal 26 ----
        elif re.search(r'PPh\s+Pasal\s+26\b', line, re.IGNORECASE):
            fill_6(extract_row_6(line), [
                "pph26_setor_sendiri", "pph26_pemotongan",
                "pph26_ditanggung_pemerintah", "pph26_harus_disetor",
                "pph26_telah_disetor", "pph26_kurang_lebih_setor",
            ])

        elif re.search(r'KJS\s*[:\s]*411127[-\s]110', line, re.IGNORECASE):
            fill_4(extract_kjs_4(line), [
                "kjs_411127_110_pemotongan", "kjs_411127_110_harus_disetor",
                "kjs_411127_110_telah_disetor", "kjs_411127_110_selisih",
            ])

        # ---- Total Pajak Penghasilan ----
        elif re.search(r'Total\s+Pajak', line, re.IGNORECASE):
            # strip leading row number and label words before collecting values
            clean = re.sub(r'^\s*\d+\s+', '', line)
            clean = re.sub(r'Total\s+Pajak.*?(?=\d)', '', clean, flags=re.IGNORECASE)
            nums = numbers_on_line(clean)
            keys = [
                "total_setor_sendiri", "total_pemotongan",
                "total_ditanggung_pemerintah", "total_harus_disetor",
                "total_telah_disetor",
            ]
            for i, k in enumerate(keys):
                if i < len(nums):
                    r[k] = nums[i]

    return r


# ── Folder batch processing ────────────────────────────────────────────────────

def process_folder(folder_path):
    folder_path = Path(folder_path)
    def _nat(p):
        import re as _re
        parts = _re.split(r"(\d+)", p.name.lower())
        return [int(x) if x.isdigit() else x for x in parts]
    pdfs = sorted(folder_path.glob("*.pdf"), key=_nat)
    if not pdfs:
        print("No PDF files found.")
        return []

    results = []
    for i, pdf in enumerate(pdfs, 1):
        print(f"[{i}/{len(pdfs)}] {pdf.name} ... ", end="", flush=True)
        t0 = time.time()
        try:
            data = process_spt_pph23(pdf)
            results.append(data)
            print(f"OK ({time.time()-t0:.1f}s)")
        except Exception as e:
            results.append({"file": pdf.name, "error": str(e)})
            print(f"ERROR: {e}")

    return results



# ── Column schema (3-row header) ───────────────────────────────────────────────
# Each entry: (row1_group, row2_kjs, row3_label, dict_key, is_numeric)
# row2_kjs = None  → cell merges rows 2+3 (INFO, TOTAL, OTHERS leaf cols)

_KJS_LEAVES = ["Pemotongan", "Harus Disetor", "Telah Disetor", "Kurang/Lebih"]

SCHEMA = []

# INFO WAJIB PAJAK
for _lbl, _key in [
    ("File",        "file"),
    ("Masa Pajak",  "masa_pajak"),
    ("Status SPT",  "status_spt"),
    ("NPWP",        "npwp"),
    ("Nama",        "nama"),
    ("No. Telepon", "no_telepon"),
]:
    SCHEMA.append(("INFO WAJIB PAJAK", None, _lbl, _key, False))

_TOTAL_LEAVES = ["Setor Sendiri", "Pemotongan", "DTP", "Harus Disetor", "Telah Disetor", "Kurang/Lebih"]

# PPh Pasal 4 ayat 2  — KJS sub-rows + Total
for _kjs, _keys in [
    ("KJS 411128-100", ["kjs_411128_100_pemotongan","kjs_411128_100_harus_disetor","kjs_411128_100_telah_disetor","kjs_411128_100_selisih"]),
    ("KJS 411128-402", ["kjs_411128_402_pemotongan","kjs_411128_402_harus_disetor","kjs_411128_402_telah_disetor","kjs_411128_402_selisih"]),
    ("KJS 411128-403", ["kjs_411128_403_pemotongan","kjs_411128_403_harus_disetor","kjs_411128_403_telah_disetor","kjs_411128_403_selisih"]),
]:
    for _leaf, _key in zip(_KJS_LEAVES, _keys):
        SCHEMA.append(("PPh Pasal 4 ayat 2", _kjs, _leaf, _key, True))
for _leaf, _key in zip(_TOTAL_LEAVES, ["pph4a2_setor_sendiri","pph4a2_pemotongan","pph4a2_ditanggung_pemerintah","pph4a2_harus_disetor","pph4a2_telah_disetor","pph4a2_kurang_lebih_setor"]):
    SCHEMA.append(("PPh Pasal 4 ayat 2", "Total", _leaf, _key, True))

# PPh Pasal 15  — KJS sub-rows + Total
for _kjs, _keys in [
    ("KJS 411128-600", ["kjs_411128_600_pemotongan","kjs_411128_600_harus_disetor","kjs_411128_600_telah_disetor","kjs_411128_600_selisih"]),
    ("KJS 411129-600", ["kjs_411129_600_pemotongan","kjs_411129_600_harus_disetor","kjs_411129_600_telah_disetor","kjs_411129_600_selisih"]),
]:
    for _leaf, _key in zip(_KJS_LEAVES, _keys):
        SCHEMA.append(("PPh Pasal 15", _kjs, _leaf, _key, True))
for _leaf, _key in zip(_TOTAL_LEAVES, ["pph15_setor_sendiri","pph15_pemotongan","pph15_ditanggung_pemerintah","pph15_harus_disetor","pph15_telah_disetor","pph15_kurang_lebih_setor"]):
    SCHEMA.append(("PPh Pasal 15", "Total", _leaf, _key, True))

# PPh Pasal 22  — KJS sub-rows + Total
for _kjs, _keys in [
    ("KJS 411122-100", ["kjs_411122_100_pemotongan","kjs_411122_100_harus_disetor","kjs_411122_100_telah_disetor","kjs_411122_100_selisih"]),
    ("KJS 411122-900", ["kjs_411122_900_pemotongan","kjs_411122_900_harus_disetor","kjs_411122_900_telah_disetor","kjs_411122_900_selisih"]),
    ("KJS 411122-910", ["kjs_411122_910_pemotongan","kjs_411122_910_harus_disetor","kjs_411122_910_telah_disetor","kjs_411122_910_selisih"]),
]:
    for _leaf, _key in zip(_KJS_LEAVES, _keys):
        SCHEMA.append(("PPh Pasal 22", _kjs, _leaf, _key, True))
for _leaf, _key in zip(_TOTAL_LEAVES, ["pph22_setor_sendiri","pph22_pemotongan","pph22_ditanggung_pemerintah","pph22_harus_disetor","pph22_telah_disetor","pph22_kurang_lebih_setor"]):
    SCHEMA.append(("PPh Pasal 22", "Total", _leaf, _key, True))

# PPh Pasal 23  — KJS sub-rows + Total
for _leaf, _key in zip(_KJS_LEAVES, ["kjs_411124_100_pemotongan","kjs_411124_100_harus_disetor","kjs_411124_100_telah_disetor","kjs_411124_100_selisih"]):
    SCHEMA.append(("PPh Pasal 23", "KJS 411124-100", _leaf, _key, True))
for _leaf, _key in zip(_TOTAL_LEAVES, ["pph23_setor_sendiri","pph23_pemotongan","pph23_ditanggung_pemerintah","pph23_harus_disetor","pph23_telah_disetor","pph23_kurang_lebih_setor"]):
    SCHEMA.append(("PPh Pasal 23", "Total", _leaf, _key, True))

# PPh Pasal 26  — KJS sub-rows + Total
for _leaf, _key in zip(_KJS_LEAVES, ["kjs_411127_110_pemotongan","kjs_411127_110_harus_disetor","kjs_411127_110_telah_disetor","kjs_411127_110_selisih"]):
    SCHEMA.append(("PPh Pasal 26", "KJS 411127-110", _leaf, _key, True))
for _leaf, _key in zip(_TOTAL_LEAVES, ["pph26_setor_sendiri","pph26_pemotongan","pph26_ditanggung_pemerintah","pph26_harus_disetor","pph26_telah_disetor","pph26_kurang_lebih_setor"]):
    SCHEMA.append(("PPh Pasal 26", "Total", _leaf, _key, True))

# Total Pajak Penghasilan
for _leaf, _key in zip(_TOTAL_LEAVES[:5], ["total_setor_sendiri","total_pemotongan","total_ditanggung_pemerintah","total_harus_disetor","total_telah_disetor"]):
    SCHEMA.append(("TOTAL", None, _leaf, _key, True))

# OTHERS
for _lbl, _key in [
    ("Nama Penandatangan", "penandatangan_nama"),
    ("Tanggal TTD",        "tanggal_tanda_tangan"),
]:
    SCHEMA.append(("OTHERS", None, _lbl, _key, False))

_GROUP_COLORS = {
    "INFO WAJIB PAJAK":   "1F4E79",
    "PPh Pasal 4 ayat 2": "375623",
    "PPh Pasal 15":       "833C00",
    "PPh Pasal 22":       "0E7C7B",
    "PPh Pasal 23":       "156082",
    "PPh Pasal 26":       "4A235A",
    "TOTAL":              "7B2C2C",
    "OTHERS":             "117A65",
}

NUMBER_FMT = '#,##0;(#,##0);"-"'


# ── Excel export ───────────────────────────────────────────────────────────────

def export_to_excel(results, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "SPT Masa PPh23"

    thin   = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(row, col, value, color, span_rows=1, span_cols=1):
        """Write a header cell and fill+border every cell in its merged range."""
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=8)
        c.fill      = PatternFill("solid", start_color=color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border
        # Apply fill+border to every cell covered by the merge so lines show correctly
        for r in range(row, row + span_rows):
            for cl in range(col, col + span_cols):
                cell = ws.cell(row=r, column=cl)
                cell.fill   = PatternFill("solid", start_color=color)
                cell.border = border
        return c

    # ── Compute column spans ───────────────────────────────────────────────────
    # (group) spans for row 1
    grp_span   = {}   # group → [first_col, last_col]
    # (group, kjs) spans for row 2
    kjs_span   = {}   # (group, kjs) → [first_col, last_col]

    for idx, (grp, kjs, lbl, key, is_num) in enumerate(SCHEMA):
        col = idx + 1
        if grp not in grp_span:
            grp_span[grp] = [col, col]
        else:
            grp_span[grp][1] = col
        tag = (grp, kjs)
        if tag not in kjs_span:
            kjs_span[tag] = [col, col]
        else:
            kjs_span[tag][1] = col

    # Row 1 — group headers
    for grp, (sc, ec) in grp_span.items():
        color  = _GROUP_COLORS.get(grp, "1F4E79")
        ncols  = ec - sc + 1
        if sc != ec:
            ws.merge_cells(f"{get_column_letter(sc)}1:{get_column_letter(ec)}1")
        _hdr(1, sc, grp, color, span_rows=1, span_cols=ncols)
    ws.row_dimensions[1].height = 22

    # Row 2 — KJS headers (or merge r2+r3 for None-kjs groups)
    for (grp, kjs), (sc, ec) in kjs_span.items():
        color = _GROUP_COLORS.get(grp, "1F4E79")
        ncols = ec - sc + 1
        if kjs is None:
            # merge rows 2-3 per column; leaf label written in row-3 loop
            for col in range(sc, ec + 1):
                ws.merge_cells(f"{get_column_letter(col)}2:{get_column_letter(col)}3")
                _hdr(2, col, None, color, span_rows=2, span_cols=1)
        else:
            if sc != ec:
                ws.merge_cells(f"{get_column_letter(sc)}2:{get_column_letter(ec)}2")
            _hdr(2, sc, kjs, color, span_rows=1, span_cols=ncols)
    ws.row_dimensions[2].height = 22

    # Row 3 — leaf labels
    for idx, (grp, kjs, lbl, key, is_num) in enumerate(SCHEMA):
        col   = idx + 1
        color = _GROUP_COLORS.get(grp, "1F4E79")
        if kjs is None:
            # overwrite value into the r2 merged cell
            ws.cell(row=2, column=col).value = lbl
        else:
            _hdr(3, col, lbl, color, span_rows=1, span_cols=1)
    ws.row_dimensions[3].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────────
    DATA_START = 4
    for row_idx, record in enumerate(results, DATA_START):
        is_error = "error" in record
        alt      = (row_idx % 2 == 0)
        for idx, (grp, kjs, lbl, key, is_num) in enumerate(SCHEMA):
            col   = idx + 1
            value = record.get(key)
            cell  = ws.cell(row=row_idx, column=col, value=value)
            cell.font   = Font(name="Arial", size=9)
            cell.border = border
            if is_num and value is not None and not is_error:
                cell.number_format = NUMBER_FMT
                cell.alignment     = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            if is_error:
                cell.fill = PatternFill("solid", start_color="FCE4D6")
            elif alt:
                cell.fill = PatternFill("solid", start_color="EBF3FB")
        ws.row_dimensions[row_idx].height = 16

    # ── Column widths ─────────────────────────────────────────────────────────
    WIDE = {"file":26,"masa_pajak":13,"status_spt":11,"npwp":17,"nama":22,
            "no_telepon":14,"penandatangan_nama":22,"tanggal_tanda_tangan":13}
    for idx, (grp, kjs, lbl, key, is_num) in enumerate(SCHEMA):
        ws.column_dimensions[get_column_letter(idx+1)].width = WIDE.get(key, 13)

    ws.freeze_panes = f"G{DATA_START}"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2  = wb.create_sheet("Ringkasan")
    thin2 = Side(style='thin', color='BFBFBF')
    b2   = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)
    for c, h in enumerate(["Keterangan","Jumlah"], 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font  = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center")
        cell.border = b2
    for ri, (label, val) in enumerate([
        ("Total file diproses", len(results)),
        ("Berhasil diekstrak",  sum(1 for r in results if "error" not in r)),
        ("Gagal diproses",      sum(1 for r in results if "error" in r)),
    ], 2):
        ws2.cell(row=ri, column=1, value=label).font = Font(name="Arial")
        ws2.cell(row=ri, column=2, value=val).font   = Font(name="Arial")
        for c in range(1, 3):
            ws2.cell(row=ri, column=c).border = b2
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14

    wb.save(output_path)
    print(f"Exported: {output_path}  ({len(results)} rows, {len(SCHEMA)} kolom)")
# ── CLI ────────────────────────────────────────────────────────────────────────

_DEFAULT_FOLDER = "PPH_23"

if __name__ == "__main__":
    args = sys.argv[1:]

    output_xlsx = None
    if "--output" in args:
        idx = args.index("--output")
        if idx + 1 < len(args):
            output_xlsx = Path(args[idx + 1])
        args = [a for i, a in enumerate(args) if i != idx and i != idx + 1]

    # Auto-detect PDF folder when no positional argument is given
    if not args:
        target = Path(_DEFAULT_FOLDER)
        if not target.is_dir():
            print(f"Error: default folder '{_DEFAULT_FOLDER}' not found. Pass a folder or file path.")
            sys.exit(1)
    else:
        target = Path(args[0])

    if target.is_dir():
        if output_xlsx is None:
            output_xlsx = Path("hasil_ekstraksi_pph23.xlsx")  # project root
        results = process_folder(target)
        if results:
            export_to_excel(results, output_xlsx)
    elif target.is_file():
        data = process_spt_pph23(target)
        if output_xlsx:
            export_to_excel([data], output_xlsx)
        else:
            for k, v in data.items():
                print(f"{k}: {v}")
    else:
        print(f"Error: '{target}' not found.")
        print(f"Usage:")
        print(f"  No args : python pph23_extractor.py              (uses ./{_DEFAULT_FOLDER}/)")
        print(f"  Folder  : python pph23_extractor.py <folder>   [--output out.xlsx]")
        print(f"  Single  : python pph23_extractor.py <file.pdf> [--output out.xlsx]")
        sys.exit(1)
        